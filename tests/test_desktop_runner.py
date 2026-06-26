from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
import pytest

from desktop_runner import NoInputFilesError, default_output_paths, resolve_input_paths, run_extraction


def make_workbook(path: Path) -> None:
    wb = Workbook()
    wb.active.title = "Other"
    wb.save(path)


def make_non_order_excel(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = "普通报表"
    ws["A2"] = "不是订单"
    ws["B2"] = "2026-06-15"
    wb.save(path)


def make_order_workbook(path: Path, job: str, delivery_date: str = "2026-06-15", po: str = "PO-1") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    ws["C1"] = job
    ws["C2"] = "Builder"
    ws["C5"] = delivery_date
    ws["C6"] = po
    headers = ["Material", "Stock", "Qty", "Profile", "B/O", "Reveal Height", "Reveal Width"]
    for col, value in enumerate(headers, start=1):
        ws.cell(9, col).value = value
    ws.cell(11, 1).value = "1.05mm Zincanneal"
    ws.cell(11, 3).value = 1
    ws.cell(11, 4).value = "Modern"
    ws.cell(11, 7).value = 823
    wb.save(path)


def test_resolve_input_paths_accepts_folder_and_filters_excel_files(tmp_path: Path) -> None:
    make_workbook(tmp_path / "order.xlsx")
    make_workbook(tmp_path / "order.xlsm")
    make_workbook(tmp_path / "~$temp.xlsx")
    (tmp_path / "notes.txt").write_text("ignore", encoding="utf-8")
    (tmp_path / "2026 Job Track.xlsx").write_text("ignore", encoding="utf-8")

    result = resolve_input_paths([tmp_path])

    assert [path.name for path in result.input_files] == ["order.xlsm", "order.xlsx"]
    assert result.base_dir == tmp_path
    assert "notes.txt" in result.skipped_files


def test_resolve_input_paths_accepts_multiple_excel_files(tmp_path: Path) -> None:
    one = tmp_path / "one.xlsx"
    two = tmp_path / "two.xlsm"
    bad = tmp_path / "bad.txt"
    make_workbook(one)
    make_workbook(two)
    bad.write_text("ignore", encoding="utf-8")

    result = resolve_input_paths([two, bad, one])

    assert [path.name for path in result.input_files] == ["one.xlsx", "two.xlsm"]
    assert result.base_dir == tmp_path
    assert result.skipped_files == ["bad.txt"]


def test_resolve_input_paths_uses_first_valid_file_parent_for_file_drops(tmp_path: Path) -> None:
    z_dir = tmp_path / "z"
    a_dir = tmp_path / "a"
    z_dir.mkdir()
    a_dir.mkdir()
    first = z_dir / "first.xlsx"
    second = a_dir / "second.xlsx"
    make_workbook(first)
    make_workbook(second)

    result = resolve_input_paths([first, second])

    assert result.base_dir == z_dir
    assert [path.name for path in result.input_files] == ["second.xlsx", "first.xlsx"]


def test_default_output_paths_use_output_folder(tmp_path: Path) -> None:
    paths = default_output_paths(tmp_path)

    assert paths.output_dir == tmp_path / "order_extraction_output"
    assert paths.xlsx_output.name == "订单整理结果.xlsx"
    assert paths.csv_output.name == "extracted_job_rows.csv"
    assert paths.audit_output.name == "audit.csv"


def test_run_extraction_rejects_no_valid_files(tmp_path: Path) -> None:
    with pytest.raises(NoInputFilesError, match="No valid order Excel files"):
        run_extraction([tmp_path / "notes.txt"])


def test_run_extraction_keeps_latest_source_version_for_duplicate_jobs(tmp_path: Path) -> None:
    older = tmp_path / "29698__0178__old__29698 BEYOND RES SPLIT + CS.xlsx"
    newer = tmp_path / "29698__0216__new__29698 BEYOND RES SPLIT + CS.xlsx"
    make_order_workbook(older, "29698")
    make_order_workbook(newer, "29698")

    result = run_extraction([tmp_path])

    assert len(result.rows) == 1
    assert result.rows[0].source_file == newer.name


def test_run_extraction_sorts_rows_by_ideal_delivery_date(tmp_path: Path) -> None:
    later = tmp_path / "100 late.xlsx"
    blank = tmp_path / "200 blank.xlsx"
    earlier = tmp_path / "300 early.xlsx"
    make_order_workbook(later, "30002", delivery_date="2026-06-20", po="LATE")
    make_order_workbook(earlier, "30001", delivery_date="2026-06-01", po="EARLY")
    make_order_workbook(blank, "30003", delivery_date="", po="BLANK")

    result = run_extraction([later, blank, earlier])

    assert [row.values[1] for row in result.rows] == ["EARLY", "LATE", "BLANK"]


def test_run_extraction_skips_excel_files_without_order_rules(tmp_path: Path) -> None:
    order = tmp_path / "29698 order.xlsx"
    report = tmp_path / "weekly report.xlsx"
    make_order_workbook(order, "29698")
    make_non_order_excel(report)

    result = run_extraction([tmp_path])

    assert len(result.rows) == 1
    assert result.rows[0].source_file == order.name
    assert report.name in result.skipped_files
